#coding=utf8
import json
import os

import openpyxl
from Clib.Threadhtml import Threadstart
from Clib.http import _http


def list_split(num,tied):
    '''平均分组
    :return:[[5, 51], [51, 101], [101, 151], [151, 201]]
    '''
    list=[]
    for i in range(2,num,tied):
        p=[i,i+tied]
        if i+tied>num:
            p=[i,num]
        list.append(p)
    return list

def list_all_dict(dict_a):
        '''
        :param dict_a: 遍历所有字段
        :return:
        '''
        keys=[]
        def list_dict(dict_a):
            if isinstance(dict_a,dict):
                for x in range(len(dict_a)):
                    key= dict_a.keys()[x]
                    value = dict_a[key]
                    list_dict(value)
                    keys.append(key)
                return keys
        return list_dict(dict_a)


class read_xsls():
    def __init__(self):
        self.readpath = os.getcwd() + '/test.xlsx'
        self.inwb = openpyxl.load_workbook(self.readpath)
        self.work_name= self.inwb.get_sheet_by_name('Sheet1')

    def high_row(self):
        high = self.work_name.get_highest_row()
        return high+1

    def high_column(self):
        high = self.work_name.get_highest_column()
        return high+1

    def get_data(self,token,url,data=None,):
        header={'User-Agent':"Mozilla/5.0 (Windows NT 6.1; WOW64; rv:26.0) Gecko/20100101 Firefox/26.0",
                'token':"%s"%token,
                'appversion':"1.8"}
        http = _http(header=header,data=data)
        self.code = http.get_code(req_url=url)
        work_data=http.get_data(req_url=url,num=3,type=2)
        return work_data


class work(read_xsls):


    def main(self,list):
        for row in range(list[0],list[1]):
            Remarks=[]
            try:
                arr={}
                arr['Name']=self.work_name.cell(row = row,column = 2).value
                arr['Url']=self.work_name.cell(row = row,column = 3).value
                arr['Token']=self.work_name.cell(row = row,column = 4).value
                arr['Data']=self.work_name.cell(row = row,column = 5).value
                arr['Code']=self.work_name.cell(row = row,column = 6).value
                arr['Example']=self.work_name.cell(row = row,column = 7).value
                arr['Method']=self.work_name.cell(row = row,column = 8).value
                if arr['Method']==u'GET':
                    url=arr['Url']+'?'+arr['Data']
                    data=None
                else:
                    url=arr['Url']
                    data=arr['Data']
                content=self.get_data(arr['Token'],url,data)

                #验证所有字段是否默认规格匹配
                Getfield=list_all_dict(content)
                p=json.loads(arr['Example'])
                xslsfield=list_all_dict(p)
                if Getfield==xslsfield:
                    pass
                else:
                    field=[]
                    for i in Getfield:
                            if i not in xslsfield:
                                field.append(u'缺失字段'+str(i))
                    Remarks=field
                    return False


                #验证返回状态
                if self.code==arr['Code']:
                    pass
                else:
                    Remarks=u'返回Url状态错误'+str(self.code+arr['Url'])
                    return False

                type="Pass"
            except Exception,error:
                Remarks.append(error)
                Remark=str(Remarks)
                type="Fail"
                self.work_name.cell(row = row,column = 10).value=Remark
            self.work_name.cell(row = row,column = 9).value=type
            self.inwb.save(self.readpath)





if __name__ == "__main__":
    work=work()
    highNum=list_split(work.high_row(),5)
    Threadstart(work.main,highNum,5)

