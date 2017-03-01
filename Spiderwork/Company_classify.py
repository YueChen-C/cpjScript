#coding=utf8
#下载xlsx导入数据库
import json
import os
import re

import openpyxl

from Clib import db
from Clib.Threadhtml import threadStart
from Clib.http import _Http


def list_split(num,tied):
    '''平均分组
    :return:[[5, 51], [51, 101], [101, 151], [151, 201]]
    '''
    list=[]
    for i in range(5,num,tied):
        p=[i,i+tied]
        if i+tied>num:
            p=[i,num]
        list.append(p)
    return list


class down():
    def down_urls(self):
        url='http://www.neeq.com.cn/info/list.do?callback=jQuery183012995692230096056_1467603134815'
        data='keywords=&page=0&pageSize=10&nodeId=259'
        http=_Http(data=data)
        content = http.getData(req_url=url, num=3)
        pattern = re.compile(r'\[.*\]', re.DOTALL).findall(content)
        htmljson = json.loads("".join(pattern))
        down_url="http://www.neeq.com.cn"+htmljson[0]['data']['content'][0]['fileUrl']
        return down_url
    def down_xsls(self):
        path=os.getcwd()
        url=self.down_urls()
        httpdown=_Http()
        return httpdown.downImage(imageUrl=url,path=path)



class read_xsls(down):
    def __init__(self):
        self.readpath=os.getcwd()+'/'+self.down_xsls()
        self.work_book = openpyxl.load_workbook(self.readpath)

    def high_row(self,content):
        high=self.work_book.get_sheet_by_name(content).get_highest_row()
        return high

    def manage_content(self,list):
        work_manage = self.work_book.get_sheet_by_name(u'管理型')
        for rx in range(list[0],list[1]):
            manage=[]
            for i in range(1,11):
                manage.append(work_manage.cell(row = rx,column = i).value)
            key=('Nasdaq','shorthand','firstcode','firstname','towcode','towname','threecode','threename','fourcode','fourname')
            db.insertOne(table='manage_classify', keys=key, values=manage, repeat=3)

    def investment_content(self,list):
        work_investment = self.work_book.get_sheet_by_name(u'投资型')
        for rx in range(list[0],list[1]):
            investment=[]
            for i in range(1,11):
                investment.append(work_investment.cell(row = rx,column = i).value)
            key=('Nasdaq','shorthand','firstcode','firstname','towcode','towname','threecode','threename','fourcode','fourname')
            db.insertOne(table='company_classify', keys=key, values=investment, repeat=3)




if __name__ == "__main__":
    read_xsls=read_xsls()
    #获取总数量平均分组多线程导入
    highNum=read_xsls.high_row(content=u'管理型')
    highNum=list_split(highNum,50)
    threadStart(read_xsls.manage_content, highNum, 10)
    threadStart(read_xsls.investment_content, highNum, 10)
